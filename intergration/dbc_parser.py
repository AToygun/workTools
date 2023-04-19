import argparse
import canmatrix
import openpyxl
from openpyxl.styles import Font
from typing import List

def load_dbc_file(dbc_file_name: str) -> canmatrix.CanMatrix:
    """从DBC文件中加载数据"""
    matrix_dict = canmatrix.formats.loadp_flat(dbc_file_name)
    return matrix_dict.frames

def get_frame_by_node(dbs: canmatrix.CanMatrix, node_name: str) -> canmatrix.CanMatrix:
    """删除非应用报文，并输出节点报文"""
    app_frame_db = canmatrix.CanMatrix()
    for frame in dbs:
        if 'yes' in frame.attributes['NmMessage'] or 'yes' in frame.attributes['NmAsrMessage']:
            continue
        elif 'yes' in frame.attributes['DiagRequest'] or 'yes' in frame.attributes['DiagResponse']:
            continue
        else:
            for transmitter in frame.transmitters:
                if node_name == transmitter:
                    app_frame_db.add_frame(frame)
            for receiver in frame.receivers:
                if node_name == receiver:
                    app_frame_db.add_frame(frame)
    return app_frame_db

def write_excel_file(app_frames: canmatrix.CanMatrix, node_name: str, excel_file_name: str):
    """将信号信息写入Excel文件"""
    # 创建一个工作簿
    workbook = openpyxl.Workbook()
    
    # 创建一个工作表，并添加信号信息
    worksheet_name = "Signal Info"
    worksheet = workbook.active
    worksheet.title = worksheet_name

    # 设置标题行样式
    bold_font = Font(bold=True)
    
    # 写入标题行
    header_row = ["ASW_Name", "ASW_Interface", "CAN_SignalName", "MsgName", "MsgID(Hex)", \
                    "Direction", "StartBit", "BitLength", "InitialValue", "Factor", \
                    "Offset", "Min", "Max", "Unit", "Comment"]
    for col, header in enumerate(header_row):
        worksheet.cell(row=1, column=col+1, value=header).font = bold_font

    # 写入每个信号的信息
    row = 2
    for frame in app_frames:
        for signal in frame.signals:
                
            message_name = frame.name
            message_id = hex(frame.arbitration_id.id)
            signal_name = signal.name
            start_bit = signal.get_startbit(bit_numbering=1)
            length = signal.size
            if node_name in frame.transmitters:
                direction = "Tx"
            else:
                direction = "Rx"
            initial_value = signal.initial_value
            factor = signal.factor
            offset = signal.offset
            minimum = signal.min
            maximum = signal.max
            unit = signal.unit
            comment = signal.comment
            
            # 写入数据行
            data_row = ["", "", signal_name, message_name, message_id, \
                        direction, start_bit, length, initial_value, factor, \
                            offset, minimum, maximum, unit, comment]
            for col, data in enumerate(data_row):
                worksheet.cell(row=row, column=col+1, value=data)
            row += 1

    # 将Excel文件写入磁盘
    workbook.save(excel_file_name)

def main():
    # 解析命令行参数
    parser = argparse.ArgumentParser(description="Convert a DBC file to an Excel file for a given node.")
    parser.add_argument("dbc_file", help="the name of the DBC file to convert")
    parser.add_argument("node_name", help="the name of the node to extract signals for")
    args = parser.parse_args()

    # 载入DBC文件
    matrix = load_dbc_file(args.dbc_file)

    # 选择节点
    app_frames = get_frame_by_node(matrix, args.node_name)

    # 创建Excel文件
    excel_file_name = f"{args.node_name}.xlsx"
    write_excel_file(app_frames, args.node_name, excel_file_name)

if __name__ == "__main__":
    main()
